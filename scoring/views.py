from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAuthenticated, AllowAny
from django.shortcuts import get_object_or_404, render
from django.utils.html import escape
from django.conf import settings
from django.views.decorators.csrf import ensure_csrf_cookie
import logging
from .models import Room, Member, Route, Score, update_scores
from .serializers import (
    RoomSerializer, MemberSerializer, RouteSerializer,
    RouteCreateSerializer, RouteUpdateSerializer, LeaderboardSerializer, ScoreUpdateSerializer
)
from .permissions import IsAuthenticatedOrReadOnlyForCreate
from .utils import get_log_file_path, get_logs_directory, get_platform_info, is_mobile_device
from .serializers import convert_file_to_uploaded_file

logger = logging.getLogger(__name__)

# Excel 导出相关导入
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as ExcelImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 未安装，Excel 导出功能将不可用。请运行: pip install openpyxl")


class RoomViewSet(viewsets.ModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    # 使用 settings.py 中的默認權限設置（開發環境為 AllowAny）
    
    def get_queryset(self):
        """確保查詢時預加載相關數據"""
        return Room.objects.prefetch_related(
            'routes__scores__member',
            'members'
        ).all()
    
    def retrieve(self, request, *args, **kwargs):
        """獲取房間詳情，確保數據最新"""
        instance = self.get_object()
        
        # 強制清除可能的緩存，重新查詢並預取相關數據
        instance.refresh_from_db()
        instance = Room.objects.prefetch_related(
            'routes__scores__member',
            'members'
        ).get(pk=instance.pk)
        
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        """創建房間"""
        data = request.data.copy()
        # 移除standard_line_score，因為它會自動計算
        if 'standard_line_score' in data:
            del data['standard_line_score']
        
        # 清理用戶輸入，防止 XSS
        if 'name' in data:
            data['name'] = escape(data['name'])
        
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            room = serializer.save()
            # 創建房間後自動計算standard_line_score
            # 注意：如果房間剛創建還沒有成員，會使用默認值12
            update_scores(room.id)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        """更新房間"""
        room = self.get_object()
        # 移除standard_line_score，因為它會自動計算
        data = request.data.copy()
        if 'standard_line_score' in data:
            del data['standard_line_score']
        
        # 清理用戶輸入，防止 XSS
        if 'name' in data:
            data['name'] = escape(data['name'])
        
        serializer = self.get_serializer(room, data=data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            # 更新後自動重新計算分數（會自動更新standard_line_score）
            update_scores(room.id)
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'], url_path='leaderboard')
    def leaderboard(self, request, pk=None):
        """獲取排行榜"""
        room = self.get_object()
        
        # 按總分降序排序
        members = room.members.all().order_by('-total_score', 'name')
        
        room_info = {
            'name': room.name,
            'standard_line_score': room.standard_line_score,
            'id': room.id
        }
        
        serializer = LeaderboardSerializer({
            'room_info': room_info,
            'leaderboard': members
        })
        
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, pk=None):
        """導出排行榜Excel，包含照片和測項"""
        if not OPENPYXL_AVAILABLE:
            return Response(
                {'detail': 'Excel 导出功能不可用，请安装 openpyxl: pip install openpyxl'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from io import BytesIO
        import os
        
        room = self.get_object()
        
        # 創建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "排行榜"
        
        # 設置標題樣式
        title_font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # 設置表頭樣式
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 設置邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入標題
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"{room.name} - 排行榜"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        title_cell.border = thin_border
        ws.row_dimensions[1].height = 30
        
        # 寫入房間信息
        ws['A2'] = f"房間 ID: {room.id}"
        ws['B2'] = f"每一條線總分 (L): {room.standard_line_score}"
        ws.merge_cells('A2:B2')
        info_cell = ws['A2']
        info_cell.font = Font(name='微軟正黑體', size=11)
        info_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 寫入表頭
        headers = ['排名', '成員', '總分', '完成條數', '是否客製化組']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[3].height = 25
        
        # 獲取成員數據（按總分降序）
        members = room.members.all().order_by('-total_score', 'name')
        
        # 寫入成員數據（處理同分排名）
        row_num = 4
        current_rank = 1
        previous_score = None
        
        for index, member in enumerate(members, start=1):
            current_score = float(member.total_score)
            
            # 第一個成員始終是第1名
            if index == 1:
                current_rank = 1
            else:
                # 如果分數與前一個不同，更新排名為當前索引
                if previous_score is not None and current_score != previous_score:
                    current_rank = index
                # 如果分數相同，保持當前排名（不更新 current_rank）
            
            previous_score = current_score
            
            ws.cell(row=row_num, column=1, value=current_rank).font = Font(name='微軟正黑體', size=11)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_num, column=1).border = thin_border
            
            ws.cell(row=row_num, column=2, value=member.name).font = Font(name='微軟正黑體', size=11)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row_num, column=2).border = thin_border
            
            ws.cell(row=row_num, column=3, value=float(member.total_score)).font = Font(name='微軟正黑體', size=11)
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=row_num, column=3).border = thin_border
            
            completed_count = member.completed_routes_count
            ws.cell(row=row_num, column=4, value=completed_count).font = Font(name='微軟正黑體', size=11)
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_num, column=4).border = thin_border
            
            custom_text = '是' if member.is_custom_calc else '否'
            ws.cell(row=row_num, column=5, value=custom_text).font = Font(name='微軟正黑體', size=11)
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_num, column=5).border = thin_border
            
            # 設置行高
            ws.row_dimensions[row_num].height = 20
            row_num += 1
        
        # 設置列寬
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # 創建路線工作表（測項）
        routes_ws = wb.create_sheet(title="路線列表")
        
        # 獲取所有成員（按名稱排序，保持一致性）
        all_members = room.members.all().order_by('name')
        member_names = [member.name for member in all_members]
        member_dict = {member.id: member for member in all_members}
        
        # 路線表頭：基本信息 + 每個成員的列
        route_headers = ['路線名稱', '難度等級', '完成人數', '照片'] + member_names
        for col_num, header in enumerate(route_headers, 1):
            cell = routes_ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        routes_ws.row_dimensions[1].height = 25
        
        # 獲取路線數據
        routes = room.routes.all().order_by('created_at')
        
        # 用於保存所有臨時文件路徑，以便最後清理
        temp_files = []
        
        # 寫入路線數據
        route_row = 2
        for route in routes:
            routes_ws.cell(row=route_row, column=1, value=route.name).font = Font(name='微軟正黑體', size=11)
            routes_ws.cell(row=route_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            routes_ws.cell(row=route_row, column=1).border = thin_border
            
            routes_ws.cell(row=route_row, column=2, value=route.grade or '-').font = Font(name='微軟正黑體', size=11)
            routes_ws.cell(row=route_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            routes_ws.cell(row=route_row, column=2).border = thin_border
            
            # 計算完成人數
            completed_count = route.scores.filter(is_completed=True).count()
            total_count = route.scores.count()
            routes_ws.cell(row=route_row, column=3, value=f"{completed_count}/{total_count}").font = Font(name='微軟正黑體', size=11)
            routes_ws.cell(row=route_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            routes_ws.cell(row=route_row, column=3).border = thin_border
            
            # 處理照片
            if route.photo:
                try:
                    photo_path = route.photo.path
                    if os.path.exists(photo_path):
                        # 調整照片大小以適應Excel單元格
                        from PIL import Image as PILImage
                        img = PILImage.open(photo_path)
                        # 限制照片大小（寬度不超過200像素）
                        max_width = 200
                        if img.width > max_width:
                            ratio = max_width / img.width
                            new_height = int(img.height * ratio)
                            img = img.resize((max_width, new_height), PILImage.Resampling.LANCZOS)
                        
                        # 保存臨時圖片
                        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_export')
                        os.makedirs(temp_dir, exist_ok=True)
                        temp_path = os.path.join(temp_dir, f'route_{route.id}.jpg')
                        img.save(temp_path, 'JPEG', quality=85)
                        
                        # 記錄臨時文件路徑，稍後清理
                        temp_files.append(temp_path)
                        
                        # 插入圖片到Excel - 優化以確保 iPhone 兼容性
                        # 關鍵：確保圖片是標準 JPEG 格式並正確嵌入（不是鏈接）
                        
                        # 確保圖片是 JPEG 格式（iPhone 兼容性最好）
                        # temp_path 已經是 JPEG（上面已保存為 JPEG），直接使用
                        excel_img = ExcelImage(temp_path)
                        
                        # 設置圖片大小（像素轉換為 Excel 單位）
                        # openpyxl 使用像素作為單位，Excel 使用點（points）
                        # 1 點 = 1/72 英寸，假設 96 DPI，1 點 ≈ 1.33 像素
                        # 所以 Excel 中的尺寸 = 像素 / 0.75
                        
                        # 限制圖片最大寬度（適應儲存格）
                        max_excel_width = 200  # Excel 單位（約 150 像素）
                        max_excel_height = 150  # Excel 單位（約 112 像素）
                        
                        # 如果圖片太大，按比例縮小
                        if excel_img.width > max_excel_width or excel_img.height > max_excel_height:
                            width_ratio = max_excel_width / excel_img.width if excel_img.width > 0 else 1
                            height_ratio = max_excel_height / excel_img.height if excel_img.height > 0 else 1
                            scale_ratio = min(width_ratio, height_ratio)
                            excel_img.width = excel_img.width * scale_ratio
                            excel_img.height = excel_img.height * scale_ratio
                        
                        # 獲取目標儲存格座標
                        col_letter = get_column_letter(4)  # D 列
                        cell_coord = f'{col_letter}{route_row}'
                        
                        # 使用 openpyxl 的標準方法插入圖片
                        # add_image 會自動將圖片嵌入到文件中（不是鏈接），這對 iPhone 兼容性很重要
                        routes_ws.add_image(excel_img, cell_coord)
                        
                        # 設置行高以適應圖片
                        # Excel 行高單位是點，1 點 ≈ 1.33 像素
                        # openpyxl 的圖片尺寸單位是像素，需要轉換
                        img_height_points = excel_img.height / 0.75 if excel_img.height > 0 else 80
                        routes_ws.row_dimensions[route_row].height = max(80, img_height_points)
                        
                        # 設置列寬以適應圖片（如果需要）
                        # Excel 列寬單位是字符寬度，需要特殊轉換
                        # 大約：列寬 = (像素寬度 / 7) + 2
                        col_width_chars = (excel_img.width / 7) + 2 if excel_img.width > 0 else 15
                        routes_ws.column_dimensions[col_letter].width = max(10, col_width_chars)
                    else:
                        routes_ws.cell(row=route_row, column=4, value='照片不存在').font = Font(name='微軟正黑體', size=10, color='999999')
                        routes_ws.cell(row=route_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
                        routes_ws.cell(row=route_row, column=4).border = thin_border
                except Exception as e:
                    logger.error(f"處理路線照片時發生錯誤: {e}")
                    routes_ws.cell(row=route_row, column=4, value='照片載入失敗').font = Font(name='微軟正黑體', size=10, color='FF0000')
                    routes_ws.cell(row=route_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
                    routes_ws.cell(row=route_row, column=4).border = thin_border
            else:
                routes_ws.cell(row=route_row, column=4, value='無照片').font = Font(name='微軟正黑體', size=10, color='999999')
                routes_ws.cell(row=route_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
                routes_ws.cell(row=route_row, column=4).border = thin_border
            
            # 添加每個成員的完成狀態（1=完成，0=未完成）
            # 獲取該路線的所有成績記錄
            route_scores = {score.member_id: score.is_completed for score in route.scores.all()}
            
            # 從第5列開始（前4列是：路線名稱、難度等級、完成人數、照片）
            member_col = 5
            for member in all_members:
                # 檢查該成員是否完成此路線
                is_completed = route_scores.get(member.id, False)
                completion_value = 1 if is_completed else 0
                
                routes_ws.cell(row=route_row, column=member_col, value=completion_value).font = Font(name='微軟正黑體', size=11)
                routes_ws.cell(row=route_row, column=member_col).alignment = Alignment(horizontal='center', vertical='center')
                routes_ws.cell(row=route_row, column=member_col).border = thin_border
                member_col += 1
            
            route_row += 1
        
        # 設置路線工作表列寬
        routes_ws.column_dimensions['A'].width = 25  # 路線名稱
        routes_ws.column_dimensions['B'].width = 15  # 難度等級
        routes_ws.column_dimensions['C'].width = 15  # 完成人數
        routes_ws.column_dimensions['D'].width = 30  # 照片
        
        # 設置成員列的寬度（從第5列開始）
        for idx, member in enumerate(all_members, start=5):
            col_letter = get_column_letter(idx)
            routes_ws.column_dimensions[col_letter].width = 10  # 每個成員列寬度設為10
        
        # 保存到內存（在保存之前，臨時文件必須存在）
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 保存完成後，清理所有臨時文件
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception as cleanup_error:
                logger.warning(f"清理臨時文件失敗 {temp_file}: {cleanup_error}")
        
        # 創建HTTP響應
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{room.name}_排行榜_{room.id}.xlsx"
        # 處理中文文件名
        from urllib.parse import quote
        response['Content-Disposition'] = f'attachment; filename="{quote(filename)}"; filename*=UTF-8\'\'{quote(filename)}'
        
        return response

    @action(detail=True, methods=['post'], url_path='routes')
    def create_route(self, request, pk=None):
        """新增路線與成績錄入"""
        room = self.get_object()
        
        # 處理 member_completions：FormData 可能將值作為列表傳遞（QueryDict），取第一個元素
        # 如果已經是字典，轉換為 JSON 字符串（因為 serializer 期望字符串）
        data = request.data.copy()
        if 'member_completions' in data:
            member_completions = data['member_completions']
            if isinstance(member_completions, list):
                member_completions = member_completions[0] if len(member_completions) > 0 else '{}'
            if isinstance(member_completions, dict):
                import json
                member_completions = json.dumps(member_completions)
            if not isinstance(member_completions, str):
                member_completions = str(member_completions)
            data['member_completions'] = member_completions
        
        # 清理用戶輸入，防止 XSS
        if 'name' in data:
            data['name'] = escape(data['name'])
        if 'grade' in data:
            data['grade'] = escape(data['grade'])
        
        serializer = RouteCreateSerializer(data=data, context={'room': room, 'request': request})
        if serializer.is_valid():
            route = serializer.save()
            
            # 強制從數據庫重新獲取 route 對象，確保 scores 關係已正確加載
            from .models import Route
            route = Route.objects.prefetch_related('scores__member').get(id=route.id)
            
            route_serializer = RouteSerializer(route, context={'request': request})
            return Response(route_serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ScoreViewSet(viewsets.ModelViewSet):
    queryset = Score.objects.all()
    serializer_class = ScoreUpdateSerializer
    # 使用 settings.py 中的默認權限設置（開發環境為 AllowAny）

    def update(self, request, *args, **kwargs):
        """更新成績狀態（標記完成/未完成）"""
        score = self.get_object()
        serializer = self.get_serializer(score, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            
            # 觸發計分更新
            update_scores(score.member.room.id)
            
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RouteViewSet(viewsets.ModelViewSet):
    queryset = Route.objects.all()
    # 使用 settings.py 中的默認權限設置（開發環境為 AllowAny）
    
    def get_serializer_class(self):
        """根據操作選擇不同的序列化器"""
        if self.action == 'create':
            return RouteCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return RouteUpdateSerializer
        return RouteSerializer
    
    def get_serializer_context(self):
        """為序列化器提供上下文"""
        context = super().get_serializer_context()
        context['request'] = self.request
        if self.action == 'create':
            # 從請求路徑中獲取房間ID（如果路徑是 /api/rooms/{room_id}/routes/）
            room_id = self.request.parser_context['kwargs'].get('room_pk')
            if room_id:
                from .models import Room
                try:
                    context['room'] = Room.objects.get(id=room_id)
                except Room.DoesNotExist:
                    pass
        return context
    
    def retrieve(self, request, *args, **kwargs):
        """獲取路線詳情"""
        route = self.get_object()
        serializer = RouteSerializer(route, context={'request': request})
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        """更新路線"""
        route_id = kwargs.get('pk')
        logger.info(f"[RouteViewSet.update] 開始更新路線 ID: {route_id}")
        
        try:
            route = self.get_object()
            logger.debug(f"[RouteViewSet.update] 獲取路線對象成功: {route.id}, 名稱: {route.name}")
            
            # 在複製 request.data 之前，先處理文件對象（避免 pickle 錯誤）
            # 如果直接使用 request.data.copy()，Django 會嘗試深拷貝，導致 BufferedRandom 無法序列化
            # 需要同時支持 JSON 格式（普通字典）和 FormData 格式（QueryDict）
            from django.http import QueryDict
            
            data = {}
            # 檢查 request.data 是 QueryDict（FormData）還是普通字典（JSON）
            is_querydict = isinstance(request.data, QueryDict)
            
            for key in request.data.keys():
                # QueryDict 的 getlist 方法可以獲取所有值（包括列表）
                # 如果是普通字典，直接獲取值
                if is_querydict:
                    values = request.data.getlist(key)
                else:
                    # JSON 格式：直接獲取值，如果是列表則保持列表，否則包裝成列表以便統一處理
                    value = request.data.get(key)
                    values = [value] if value is not None else []
                
                # 如果是文件對象（photo 字段），只取第一個值並轉換
                if key == 'photo' and values:
                    photo_value = values[0] if values else None
                    if photo_value:
                        logger.debug(f"[RouteViewSet.update] 檢測到照片文件，原始類型: {type(photo_value)}")
                        try:
                            # 轉換文件對象為 InMemoryUploadedFile
                            converted_photo = convert_file_to_uploaded_file(photo_value)
                            logger.debug(f"[RouteViewSet.update] 照片文件已轉換，新類型: {type(converted_photo)}")
                            data[key] = converted_photo
                        except Exception as e:
                            logger.error(f"[RouteViewSet.update] 轉換照片文件時發生錯誤: {e}")
                            # 如果轉換失敗，仍然嘗試使用原文件（可能會導致錯誤，但至少記錄了錯誤）
                            data[key] = photo_value
                    else:
                        data[key] = None
                else:
                    # 非文件字段，處理列表值（QueryDict 可能返回列表）
                    if len(values) == 1:
                        data[key] = values[0]
                    else:
                        data[key] = values
            
            logger.debug(f"[RouteViewSet.update] 接收到的數據字段: {list(data.keys())}")
            
            # 清理用戶輸入，防止 XSS
            if 'name' in data:
                original_name = data['name']
                # 確保 name 是字符串（如果是列表，取第一個元素）
                if isinstance(original_name, list):
                    original_name = original_name[0] if len(original_name) > 0 else ''
                if isinstance(original_name, str):
                    data['name'] = escape(original_name)
                else:
                    data['name'] = str(original_name)
                logger.debug(f"[RouteViewSet.update] 清理路線名稱: '{original_name}' -> '{data['name']}'")
            if 'grade' in data:
                original_grade = data['grade']
                # 確保 grade 是字符串（如果是列表，取第一個元素）
                if isinstance(original_grade, list):
                    original_grade = original_grade[0] if len(original_grade) > 0 else ''
                if isinstance(original_grade, str):
                    data['grade'] = escape(original_grade)
                else:
                    data['grade'] = str(original_grade)
                logger.debug(f"[RouteViewSet.update] 清理難度等級: '{original_grade}' -> '{data['grade']}'")
            
            # 處理 member_completions：FormData 可能將值作為列表傳遞（QueryDict），取第一個元素
            # 如果已經是字典，轉換為 JSON 字符串（因為 serializer 期望字符串）
            if 'member_completions' in data:
                member_completions = data['member_completions']
                logger.debug(f"[RouteViewSet.update] member_completions 原始類型: {type(member_completions)}, 值: {member_completions}")
                
                if isinstance(member_completions, list):
                    member_completions = member_completions[0] if len(member_completions) > 0 else '{}'
                    logger.debug(f"[RouteViewSet.update] member_completions 從列表提取: {member_completions}")
                if isinstance(member_completions, dict):
                    import json
                    member_completions = json.dumps(member_completions)
                    logger.debug(f"[RouteViewSet.update] member_completions 從字典轉換為 JSON: {member_completions}")
                if not isinstance(member_completions, str):
                    member_completions = str(member_completions)
                    logger.debug(f"[RouteViewSet.update] member_completions 轉換為字符串: {member_completions}")
                data['member_completions'] = member_completions
            
            # 處理 photo_url：如果存在但為空或無效，移除它（避免 URLField 驗證錯誤）
            # 注意：如果前端沒有發送 photo_url，則不處理（保持現有值）
            if 'photo_url' in data:
                photo_url = data['photo_url']
                logger.debug(f"[RouteViewSet.update] photo_url 原始值類型: {type(photo_url)}, 值: {photo_url}")
                
                # 如果是列表（FormData 可能將值作為列表），取第一個元素
                if isinstance(photo_url, list):
                    photo_url = photo_url[0] if len(photo_url) > 0 else ''
                    logger.debug(f"[RouteViewSet.update] photo_url 從列表提取: {photo_url}")
                
                # 如果是空字符串或無效值，移除該字段（不更新 photo_url）
                if not photo_url or (isinstance(photo_url, str) and photo_url.strip() == ''):
                    logger.warning(f"[RouteViewSet.update] photo_url 為空或無效，移除該字段以避免驗證錯誤")
                    data.pop('photo_url', None)
                # 如果是字符串但不符合基本 URL 格式，也移除
                elif isinstance(photo_url, str):
                    photo_url = photo_url.strip()
                    if not (photo_url.startswith('http://') or photo_url.startswith('https://') or 
                            photo_url.startswith('/') or photo_url.startswith('data:')):
                        logger.warning(f"[RouteViewSet.update] photo_url 不符合 URL 格式: '{photo_url}'，移除該字段")
                        data.pop('photo_url', None)
                    else:
                        logger.debug(f"[RouteViewSet.update] photo_url 驗證通過: {photo_url}")
            
            # 處理照片文件
            if 'photo' in data:
                photo = data['photo']
                if photo:
                    logger.debug(f"[RouteViewSet.update] 收到照片文件: 名稱={getattr(photo, 'name', 'N/A')}, "
                                f"大小={getattr(photo, 'size', 'N/A')}, "
                                f"content_type={getattr(photo, 'content_type', 'N/A')}")
                else:
                    logger.debug(f"[RouteViewSet.update] photo 字段存在但為空")
            
            logger.debug(f"[RouteViewSet.update] 準備傳遞給 serializer 的數據字段: {list(data.keys())}")
            
            serializer = self.get_serializer(route, data=data, partial=True, context={'request': request})
            
            logger.debug(f"[RouteViewSet.update] 開始驗證數據...")
            if not serializer.is_valid():
                logger.error(f"[RouteViewSet.update] 數據驗證失敗！路線 ID: {route_id}")
                logger.error(f"[RouteViewSet.update] 驗證錯誤詳情: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            logger.debug(f"[RouteViewSet.update] 數據驗證通過，開始保存...")
            serializer.save()
            logger.info(f"[RouteViewSet.update] 路線更新成功: {route_id}")
            
            # 重新從數據庫獲取路線，確保包含最新的 scores
            route.refresh_from_db()
            route_serializer = RouteSerializer(route, context={'request': request})
            logger.debug(f"[RouteViewSet.update] 返回更新後的路線數據")
            return Response(route_serializer.data)
            
        except Exception as e:
            import traceback
            import sys
            exc_type, exc_value, exc_traceback = sys.exc_info()
            
            logger.exception(f"[RouteViewSet.update] 更新路線時發生未預期的錯誤！路線 ID: {route_id}")
            logger.error(f"[RouteViewSet.update] 錯誤類型: {type(e).__name__}")
            logger.error(f"[RouteViewSet.update] 錯誤信息: {str(e)}")
            logger.error(f"[RouteViewSet.update] 錯誤模塊: {type(e).__module__}")
            logger.error(f"[RouteViewSet.update] 錯誤文件: {exc_traceback.tb_frame.f_code.co_filename if exc_traceback else 'N/A'}")
            logger.error(f"[RouteViewSet.update] 錯誤行號: {exc_traceback.tb_lineno if exc_traceback else 'N/A'}")
            
            # 記錄完整的堆棧跟踪
            full_traceback = traceback.format_exception(exc_type, exc_value, exc_traceback)
            logger.error(f"[RouteViewSet.update] 完整錯誤堆棧:\n{''.join(full_traceback)}")
            
            # 記錄所有堆棧幀
            if exc_traceback:
                logger.error(f"[RouteViewSet.update] 堆棧幀詳情:")
                frame = exc_traceback
                frame_num = 0
                while frame:
                    logger.error(f"  幀 {frame_num}: {frame.tb_frame.f_code.co_filename}:{frame.tb_lineno} in {frame.tb_frame.f_code.co_name}")
                    # 記錄局部變量（如果可能）
                    try:
                        local_vars = frame.tb_frame.f_locals
                        if 'photo' in local_vars:
                            photo_obj = local_vars['photo']
                            logger.error(f"    局部變量 'photo': 類型={type(photo_obj)}, 名稱={getattr(photo_obj, 'name', 'N/A')}")
                        if 'validated_data' in local_vars:
                            logger.error(f"    局部變量 'validated_data': 鍵={list(local_vars['validated_data'].keys())}")
                    except:
                        pass
                    frame = frame.tb_next
                    frame_num += 1
                    if frame_num > 50:  # 限制最多50個幀
                        logger.error(f"  ... (還有更多幀)")
                        break
            
            # 檢查是否為 pickle 錯誤
            if 'pickle' in str(e).lower() or 'BufferedRandom' in str(e) or 'BufferedReader' in str(e):
                logger.critical(f"[RouteViewSet.update] ⚠️ 檢測到 PICKLE 錯誤！")
                logger.critical(f"[RouteViewSet.update] 這通常發生在嘗試序列化 BufferedRandom 或 BufferedReader 對象時")
                logger.critical(f"[RouteViewSet.update] 請檢查文件對象是否在保存前被正確轉換為 InMemoryUploadedFile")
            
            return Response(
                {'detail': f'更新路線時發生錯誤: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def destroy(self, request, *args, **kwargs):
        """刪除路線"""
        route = self.get_object()
        room_id = route.room.id
        
        # 刪除路線（會級聯刪除相關的 Score 記錄）
        route.delete()
        
        # 觸發計分更新
        update_scores(room_id)
        
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'], url_path='log-info')
    def log_info(self, request):
        """
        獲取日誌文件信息（用於移動設備調試）
        返回日誌文件位置和平台信息
        """
        try:
            log_file_path = get_log_file_path()
            logs_dir = get_logs_directory()
            
            # 檢查日誌文件是否存在
            from pathlib import Path
            log_file = Path(log_file_path)
            file_exists = log_file.exists()
            file_size = log_file.stat().st_size if file_exists else 0
            
            # 獲取平台信息
            platform_info = get_platform_info()
            
            return Response({
                'log_file_path': log_file_path,
                'logs_directory': str(logs_dir),
                'file_exists': file_exists,
                'file_size': file_size,
                'file_size_mb': round(file_size / (1024 * 1024), 2) if file_size > 0 else 0,
                'is_mobile': is_mobile_device(),
                'platform_info': platform_info,
                'message': '日誌文件已配置。在移動設備上，日誌文件會自動保存到應用數據目錄。'
            })
        except Exception as e:
            logger.exception(f"[RouteViewSet.log_info] 獲取日誌信息時發生錯誤")
            return Response({
                'error': str(e),
                'message': '無法獲取日誌文件信息'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MemberViewSet(viewsets.ModelViewSet):
    queryset = Member.objects.all()
    serializer_class = MemberSerializer
    # 使用 settings.py 中的默認權限設置（開發環境為 AllowAny）

    def create(self, request, *args, **kwargs):
        """創建成員"""
        data = request.data.copy()
        # 清理用戶輸入，防止 XSS
        if 'name' in data:
            data['name'] = escape(data['name'])
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            member = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def update(self, request, *args, **kwargs):
        """更新成員"""
        member = self.get_object()
        data = request.data.copy()
        # 清理用戶輸入，防止 XSS
        if 'name' in data:
            data['name'] = escape(data['name'])
        serializer = self.get_serializer(member, data=data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            # 更新後自動重新計算分數（會自動更新standard_line_score）
            update_scores(member.room.id)
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        """刪除成員"""
        try:
            member = self.get_object()
        except Exception as e:
            # 如果成員不存在，返回 404
            return Response(
                {'detail': '找不到指定的成員', 'error': str(e)},
                status=status.HTTP_404_NOT_FOUND
            )
        
        room_id = member.room.id
        
        try:
            # 刪除成員（會級聯刪除相關的 Score 記錄）
            member.delete()
            
            # 觸發計分更新（會自動更新standard_line_score）
            update_scores(room_id)
            
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            # 刪除失敗，返回錯誤信息
            return Response(
                {'detail': '刪除成員時發生錯誤', 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='completed-routes')
    def completed_routes(self, request, pk=None):
        """獲取成員完成的路線列表"""
        member = self.get_object()
        
        # 獲取該成員所有完成的路線
        completed_scores = member.scores.filter(is_completed=True).select_related('route')
        routes = [score.route for score in completed_scores]
        
        # 使用 RouteSerializer 序列化路線
        from .serializers import RouteSerializer
        route_serializer = RouteSerializer(routes, many=True, context={'request': request})
        
        return Response({
            'member_id': member.id,
            'member_name': member.name,
            'completed_routes': route_serializer.data,
            'total_count': len(routes)
        })


@ensure_csrf_cookie
def index_view(request):
    """首頁視圖 - 未登錄顯示登錄界面，已登錄顯示房間列表"""
    return render(request, 'index.html', {
        'user': request.user if request.user.is_authenticated else None
    })


@ensure_csrf_cookie
def leaderboard_view(request, room_id):
    """排行榜頁面視圖"""
    return render(request, 'leaderboard.html', {'room_id': room_id})


@ensure_csrf_cookie
def rules_view(request):
    """說明頁面視圖"""
    return render(request, 'rules.html')

